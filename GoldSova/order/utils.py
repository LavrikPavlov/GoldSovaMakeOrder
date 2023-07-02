from openpyxl import load_workbook
from datetime import date 

def work_with_excel(file, adress):
    wb = load_workbook(file)
    ws = wb.active
    ws['A1'].value = adress
    a = ws['A1'].value
    print(f'Адресс пользователя: {a}')
    wb.save(file)


def user_directory_path(instance, filename):
    string = date.today()
    return 'orders/{0}/{1}/{2}'.format(instance.email.email, string, filename)